"""Export implementation plan tasks to .xlsx with waterfall chart and timeline."""
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).with_name("implementation-tasks.xlsx")
LIDOCHKA_OUT = Path(__file__).with_name("PG Queue Balancer — задачи (Лидочка).xlsx")

# Старт проекта (понедельник) — для календарных дат на листе «Сроки»
PROJECT_START = date(2026, 6, 9)

blocks_meta = {
    "Z": ("Подготовка и контракты", 0, "S1/S2"),
    "A": ("Схема PostgreSQL", 1, "S1"),
    "B": ("Репозитории", 1, "S1"),
    "C": ("Worker-Диспетчер", 2, "S2"),
    "D": ("Адаптер + SessionClump", 3, "S2/S3"),
    "E": ("Ошибки и идемпотентность", 4, "S4"),
    "F": ("Продюсеры", 5, "S5"),
    "G": ("Мониторинг и cleanup", "5-6", "S5/S6"),
}

tasks = [
    ("Z", "Z1", "ADR: слои PG / Worker / Adapter / SessionClump", "Единое понимание «что куда»; запрет логики очереди внутри session_registry", "—", "Да", "Документ согласован, пути из §3 плана"),
    ("Z", "Z2", "Решение по таблице каналов: source_channels vs staging queue_channels", "ТЗ требует поля assigned_account_id, extra_data_collected; в discovery только JSON clump", "—", "Да", "Запись в ADR + миграционный путь"),
    ("Z", "Z3", "Feature flags в .env.example", "Постепенный rollout без регрессий; diff §15 #35", "Z1", "Да", "Все флаги описаны с дефолтами"),
    ("Z", "Z4", "Docker-compose: опциональный сервис queue-worker", "Worker не в FastAPI lifespan — отдельный процесс по ТЗ §29", "A2", "Да", "docker compose up queue-worker стартует"),
    ("A", "A1", "QUEUE_DATABASE_URL + .env.example", "Отдельная БД очереди; не смешивать с SQLite discovery", "Z3", "Да", "Подключение из теста"),
    ("A", "A2", "async PG driver в requirements.txt", "Async worker + FastAPI без блокировки event loop", "A1", "Да", "pip install в CI"),
    ("A", "A3", "Миграция task_types", "ТЗ §8: лимиты, приоритеты, retry — в БД, не в ClumpConfig/env", "A2", "Да", "Таблица + downgrade"),
    ("A", "A4", "Миграция accounts", "ТЗ §6: hourly_limit, cooldown_until, current_task_id; diff #2, #32", "A2", "Да", "UNIQUE на session_name"),
    ("A", "A5", "Миграция task_queue + индексы + partial unique dedup_key", "ТЗ §9; diff #2, #13, #33 — атомарная задача, защита от дублей", "A3", "Да", "Индекс диспетчера + dedup constraint"),
    ("A", "A6", "Миграция account_resource_usage", "ТЗ §11; diff #5, #19 — попытки/час, включая неуспешные", "A4", "Да", "Индекс (account_id, created_at)"),
    ("A", "A7", "Миграция task_attempts", "ТЗ §10; diff #4 — диагностика, связь с attempt_count", "A5", "Да", "FK на task_queue, accounts"),
    ("A", "A8", "Расширить source_channels или создать queue_channels", "ТЗ §5.1, §22; diff #6 — продюсеры и dual-write каналов", "A4, Z2", "Да", "Поля assigned_account_id, extra_data_collected, last_updated_at"),
    ("A", "A9", "Seed task_types: 4 типа, 2 enabled на MVP", "Сразу видны приоритеты: add=500, move=100; update/collect выключены", "A3", "Да", "parser_add_channel, move_channel enabled"),
    ("A", "A10", "Скрипт sync accounts из SQLite account_store + clump sessions", "Мост discovery ↔ PG; admin_blocked → disabled", "A4", "Да", "Идемпотентный upsert, повторный запуск безопасен"),
    ("A", "A11", "Миграционный runner / Makefile target migrate-queue", "Воспроизводимое развёртывание на staging/prod", "A3–A8", "Да", "Одна команда применяет все миграции"),
    ("B", "B1", "queue/db.py: pool, транзакции, healthcheck", "Единая точка подключения; GET /health может проверять PG", "A2", "Да", "Pool + SELECT 1"),
    ("B", "B2", "repos/task_types.py", "Чтение min_available_resource_percent, retry-параметров при dispatch", "B1, A3", "Да", "get_by_code, list_enabled"),
    ("B", "B3", "repos/task_queue.py: enqueue + dedup", "API/продюсеры создают задачи; diff #33 — нет дублей активных", "B1, A5", "Да", "Конфликт dedup → skip или 409"),
    ("B", "B4", "repos/task_queue.py: claim_next (SKIP LOCKED)", "ТЗ §14; diff #7 — два воркера не берут одну задачу", "B3", "Да", "Тест: 2 claim → разные id"),
    ("B", "B5", "repos/task_queue.py: complete, fail, retry, postpone", "ТЗ §13; diff #16 — postpone_count без расхода ресурса", "B4", "Да", "Каждый переход статуса покрыт"),
    ("B", "B6", "repos/accounts.py: pick, reserve, release", "ТЗ §16; diff #32 — current_task_id на время execute", "B1, A4", "Да", "reserve атомарен в транзакции"),
    ("B", "B7", "repos/accounts.py: cooldown, banned, sync admin_blocked", "Связка runtime health ↔ PG; diff #37", "B6", "Да", "set_cooldown после flood из теста"),
    ("B", "B8", "repos/resource_usage.py: insert, count_last_hour", "Формула available_resource из ТЗ §7", "B1, A6", "Да", "count совпадает с insert за час"),
    ("B", "B9", "repos/task_attempts.py: insert, finish", "История для мониторинга и отладки retry", "B1, A7", "Да", "attempt_number монотонен"),
    ("B", "B10", "Unit-тесты репозиториев", "Регрессии на claim/dedup/reserve — критичные инварианты", "B3–B9", "Да", "CI green без реального Telethon"),
    ("C", "C1", "queue_worker.py: asyncio loop, SIGTERM shutdown", "Отдельный процесс; graceful stop без зависших in_progress", "B4", "Да", "После SIGTERM задачи доходят до release"),
    ("C", "C2", "Цикл claim → reserve → adapter.execute → release", "Ядро ТЗ §5.8, §13", "C1, B4, B6, D3", "Да", "Mock adapter: queued → done"),
    ("C", "C3", "Postpone при нет аккаунта (+5 min, postpone_count++)", "ТЗ §15.3, §16; diff #34 — следующая задача обрабатывается", "C2, B5", "Да", "2 задачи: вторая done пока первая scheduled"),
    ("C", "C4", "Dual reserve для move_channel (source + target)", "ТЗ §18–19; diff #12, #21", "C2, B6", "Да", "Оба current_task_id в одной TX"),
    ("C", "C5", "Проверка min_available_resource_percent", "ТЗ §8, §16; diff #20 — не брать аккаунт «на грани»", "C2, B8, B2", "Да", "Задача postpone при 79% при пороге 80%"),
    ("C", "C6", "Watchdog: in_progress → stuck по task_timeout_seconds", "ТЗ §13.4; diff #29", "C1, B5, B2", "Да", "Зависший mock → stuck + лог"),
    ("C", "C7", "Сортировка claim: priority DESC, created_at ASC", "ТЗ §15.2; diff #17", "B4", "Да", "HIGH priority задача раньше LOW"),
    ("C", "C8", "Mock-adapter integration test", "Сквозной dispatch без Telethon", "C2", "Да", "pytest в CI"),
    ("C", "C9", "Multi-worker test (2 процесса, SKIP LOCKED)", "ТЗ §14; diff #7", "C4", "Да", "N задач → N воркеров, без дублей"),
    ("D", "D1", "SessionClump.add_channel_on_session(session, ref, webhook)", "Worker уже выбрал аккаунт; обход _pick_target", "—", "Да", "Канал на указанной сессии, listener стартует"),
    ("D", "D2", "SessionClump.move_channel(ref, from, to)", "Один канал для задачи move_channel; идемпотентнее batch migrate", "—", "Да", "assignments обновлён, source очищен"),
    ("D", "D3", "queue/adapter.py: execute_task → parser_add_channel", "Маппинг task_type → Telethon; заполнение queue_prot.py re-export", "D1, C2", "Да", "Payload из §1.3 плана отрабатывает"),
    ("D", "D4", "queue/adapter.py: ветка move_channel", "Продюсер ±5% и аварийный перенос через одну задачу", "D2, C4", "Да", "source+target из payload"),
    ("D", "D5", "INSERT account_resource_usage при старте execute", "ТЗ §7.3; diff #19 — расход в момент передачи аккаунту", "D3, B8", "Да", "Запись до Telethon RPC"),
    ("D", "D6", "HealthMonitor → PG: flood → cooldown, ban → banned", "In-memory health не переживает рестарт; PG — для worker", "B7", "Да", "После mark_flood аккаунт не pick"),
    ("D", "D7", "Dual-write: PG assigned_account_id + JSON clump", "Diff #6 — два источника; пока нет полного cutover", "D3, A8", "Да", "PG и clump согласованы после add"),
    ("D", "D8", "USE_PG_QUEUE: async add-channels → N INSERT в task_queue", "Diff #18 — 1 bulk action → N атомарных задач", "B3, D3, Z3", "Да", "Flag on: action_id → task_ids[]"),
    ("D", "D9", "USE_PG_QUEUE: async remove-channels → N задач parser_remove_channel", "Симметрия add; иначе половина API на старом SQLite", "B3, D3", "Нет", "Тип задачи в seed + adapter ветка"),
    ("D", "D10", "GET /queue/tasks/{id} или reuse action API shape", "Клиент/n8n ждёт статус async операции", "B3, D8", "Да", "status, attempt_count, last_error"),
    ("D", "D11", "queue_prot.py: re-export execute_task", "Пустой файл — заготовка; единая точка импорта", "D3", "Да", "from discovery_api.queue_prot import ..."),
    ("D", "D12", "E2E: API → PG → worker → clump (staging)", "Критерий MVP §8", "D8, C2", "Да", "Реальный канал на staging"),
    ("E", "E1", "Typed errors в адаптере: RetryableError, PermanentError, ResourceError", "Worker знает, что делать без парсинга строк", "D3", "Нет", "Иерархия исключений"),
    ("E", "E2", "Маппинг classify_telethon_error → действие worker", "Переиспользуем session_health.py; diff §14", "E1, D6", "Нет", "flood → retry+cooldown"),
    ("E", "E3", "Retry с backoff из task_types", "ТЗ §20; diff #14", "E2, B5", "Нет", "2-я попытка с задержкой 10→20s"),
    ("E", "E4", "task_attempts на каждую реальную попытку", "ТЗ §10; связь attempt_count ↔ история", "D3, B9", "Нет", "1 attempt = 1 execute"),
    ("E", "E5", "Стабильные last_error коды", "Мониторинг §27; diff #29", "E1, B5", "Нет", "Коды в доке runbook"),
    ("E", "E6", "Идемпотентность: payload.last_completed_step", "ТЗ §29; collect_extra_data — пайплайн ops из ops_catalog.py", "E3", "Нет", "Retry пропускает шаги"),
    ("E", "E7", "Каталог op-кодов из queue_prot_blance/ops_catalog.py", "Гранулярный RPH per-op — следующий шаг после MVP", "E6", "Нет", "op_code в account_resource_usage"),
    ("E", "E8", "Тест: retry продолжает с упавшего op", "Приёмка идемпотентности", "E6", "Нет", "pytest green"),
    ("F", "F1", "producers/base.py: dedup + target_queue_size", "ТЗ §8.3, §12 — не переполнить очередь, не дублировать", "B3, B2", "Нет", "helper enqueue_if_room"),
    ("F", "F2", "channel_balancer.py: ±5%, INSERT move_channel", "ТЗ §22; diff #24, #26 vs rebalance_idle", "F1, A8", "Нет", "Задачи только при skew >5%"),
    ("F", "F3", "Env: REBALANCE_IDLE_ENABLED=false при PG balancer", "Два механизма переноса конфликтуют (diff §11)", "F2, Z3", "Нет", "Документировано в runbook"),
    ("F", "F4", "collect_extra_data.py", "ТЗ §23; diff #27", "F1, A8", "Нет", "До target_queue_size задач"),
    ("F", "F5", "update_channel.py", "ТЗ §24; diff #28", "F1, A8", "Нет", "Старые каналы по last_updated_at"),
    ("F", "F6", "Adapter: collect_extra_data (scorer / messages)", "Исполнение задач продюсера", "F4, D3", "Нет", "extra_data_collected=true"),
    ("F", "F7", "Adapter: update_channel", "Исполнение обновления метаданных", "F5, D3", "Нет", "last_updated_at обновлён"),
    ("F", "F8", "Cron / docker schedule для продюсеров", "Задачи не создаются сами", "F2, F4, F5", "Нет", "3 job в compose"),
    ("F", "F9", "Включить is_enabled для update/collect в seed", "После готовности adapter-веток", "F6, F7", "Нет", "Migration seed update"),
    ("G", "G1", "SQL views: queue_size_by_status, oldest_queued_age", "ТЗ §26.2; diff #29", "A5", "Нет", "View в миграции"),
    ("G", "G2", "SQL views: resource %, accounts_in_cooldown", "ТЗ §26.3", "A6, A4", "Нет", "account_available_resource_percent"),
    ("G", "G3", "GET /queue/metrics (admin API)", "Единая точка для админки/n8n", "G1, G2", "Нет", "JSON как в ТЗ"),
    ("G", "G4", "Алерты: high postpone_count, queue growth, no free accounts", "ТЗ §26.4; diff #31", "G3", "Нет", "Webhook или лог ERROR"),
    ("G", "G5", "Watchdog stuck → опционально auto-retry", "ТЗ §13.4 — не только маркировка", "C6, E3", "Нет", "Политика в ADR"),
    ("G", "G6", "Удалить action_queue.py, worker в main.py", "Diff #2 — один контур очереди", "D12", "Нет", "Нет импортов action_queue"),
    ("G", "G7", "Тесты test_action_queue.py → PG", "Регрессии bulk async", "G6", "Нет", "Новый test_queue_task.py"),
    ("G", "G8", "Sync bulk add/remove → только PG", "Diff #18, #35 — убрать _pick_target с HTTP bulk", "G6", "Нет", "USE_PG_QUEUE default true"),
    ("G", "G9", "Deprecate _add_timestamps (QUEUE_GRANULAR_RPH / PG only)", "Diff #5, #19 — один учёт RPH", "E7, B8", "Нет", "Flag off → нет in-memory счётчика"),
    ("G", "G10", "Документация runbook: worker, producers, flags", "Онбординг и incident response", "G8", "Нет", "USAGE.md или docs/queue-runbook.md"),
    ("G", "G11", "Приёмочный чеклист §8 плана (MVP + полное ТЗ)", "Формальное закрытие", "G3–G10", "Нет", "Все галочки"),
]

# Простое описание для столбца «Функция» — без имён модулей, таблиц и переменных
TASK_PLAIN_DESC: dict[str, str] = {
    "Z1": "Зафиксировать на бумаге, какие части системы за что отвечают: база задач, фоновый обработчик, мост к Telegram и управление сессиями. Не смешивать логику очереди с управлением сессиями.",
    "Z2": "Решить, где хранить список каналов и какие поля нужны: кому назначен канал, собраны ли дополнительные данные. Записать решение и путь миграции.",
    "Z3": "Добавить переключатели в конфигурацию, чтобы поэтапно включать новую очередь без поломки текущего поведения.",
    "Z4": "Добавить в docker-compose отдельный сервис фонового обработчика очереди — не внутри веб-API.",
    "A1": "Настроить отдельное подключение к базе PostgreSQL только для очереди задач, не смешивая с локальной базой discovery.",
    "A2": "Подключить асинхронный драйвер PostgreSQL в зависимости проекта.",
    "A3": "Создать в базе справочник типов задач: лимиты, приоритеты и правила повторных попыток хранятся в БД, а не в конфиге.",
    "A4": "Создать таблицу аккаунтов Telegram: почасовой лимит, пауза после flood, привязка к текущей задаче, уникальность по имени сессии.",
    "A5": "Создать таблицу очереди задач с защитой от дубликатов и индексами для быстрой выдачи воркеру.",
    "A6": "Создать таблицу учёта обращений к Telegram с каждого аккаунта за последний час (включая неуспешные).",
    "A7": "Создать таблицу истории попыток выполнения задач для диагностики и мониторинга.",
    "A8": "Расширить или добавить таблицу каналов: кому назначен, собраны ли доп. данные, когда обновлялись метаданные.",
    "A9": "Заполнить справочник типов задач: на MVP включены добавление и перенос канала; обновление и сбор данных — выключены.",
    "A10": "Скрипт синхронизации аккаунтов из локального хранилища discovery в PostgreSQL; заблокированные админом — отключены.",
    "A11": "Одна команда для применения всех миграций очереди на staging и production.",
    "B1": "Единый модуль подключения к базе очереди: пул соединений, транзакции, проверка «жива ли база».",
    "B2": "Слой доступа к настройкам типов задач: минимальный остаток ресурса, параметры повторов.",
    "B3": "Постановка задачи в очередь: API и продюсеры создают задачи; активные дубликаты не допускаются.",
    "B4": "Атомарная выдача следующей задачи воркеру — два воркера не возьмут одну и ту же задачу.",
    "B5": "Переводы статуса задачи: успех, ошибка, повтор, отложить на потом без списания лимита.",
    "B6": "Выбор, резервирование и освобождение аккаунта под задачу; на время работы аккаунт занят одной задачей.",
    "B7": "Установка паузы после flood, блокировки и синхронизация с админской блокировкой аккаунта.",
    "B8": "Запись и подсчёт использования ресурса аккаунта за последний час по формуле из ТЗ.",
    "B9": "Запись каждой попытки выполнения задачи с порядковым номером.",
    "B10": "Автотесты слоя доступа к БД на критичные сценарии без реального Telegram.",
    "C1": "Отдельный фоновый процесс-воркер с корректной остановкой по сигналу завершения.",
    "C2": "Основной цикл воркера: взять задачу → зарезервировать аккаунт → выполнить через адаптер → освободить аккаунт.",
    "C3": "Если нет свободного аккаунта — отложить задачу на 5 минут и обрабатывать следующие, не блокируя очередь.",
    "C4": "При переносе канала одновременно резервировать исходный и целевой аккаунты в одной транзакции.",
    "C5": "Не назначать аккаунт, у которого осталось слишком мало свободного лимита запросов к Telegram.",
    "C6": "Находить задачи, зависшие в статусе «в работе», и помечать их как застрявшие по таймауту.",
    "C7": "Сначала обрабатывать задачи с высоким приоритетом; при равном приоритете — более ранние.",
    "C8": "Интеграционный тест полного цикла обработки с заглушкой вместо реального Telegram.",
    "C9": "Тест двух воркеров параллельно: каждая задача обрабатывается ровно один раз.",
    "D1": "Добавлять канал на конкретный аккаунт, уже выбранный воркером, без повторного автовыбора аккаунта.",
    "D2": "Переносить один канал с одного аккаунта на другой; обновлять назначения и очищать источник.",
    "D3": "Выполнять задачу «добавить канал в парсер» через Telethon по данным из очереди.",
    "D4": "Выполнять задачу «перенести канал» между аккаунтами по данным из очереди.",
    "D5": "Списывать лимит ресурса аккаунта в момент начала работы — до любых вызовов Telegram.",
    "D6": "При flood или бане в runtime обновлять состояние аккаунта в PostgreSQL, чтобы воркер его не брал после рестарта.",
    "D7": "Пока идёт переход: дублировать назначение канала и в базе очереди, и в JSON-кластере сессий.",
    "D8": "При включённом режиме PG: массовое добавление каналов через API создаёт отдельную задачу на каждый канал.",
    "D9": "Массовое удаление каналов тоже через очередь — отдельная задача на каждый канал.",
    "D10": "HTTP-эндпоинт статуса задачи для клиента и n8n: статус, число попыток, последняя ошибка.",
    "D11": "Единая точка импорта для выполнения задач очереди из остального кода discovery.",
    "D12": "Сквозной тест на staging: запрос API → база → воркер → реальный канал в парсере.",
    "E1": "В адаптере различать ошибки: повторить позже, завершить навсегда, нет ресурса — без разбора текста ошибки.",
    "E2": "Связать классификацию ошибок Telegram с действиями воркера: flood → пауза и повтор, и т.д.",
    "E3": "Повторные попытки с нарастающей задержкой по настройкам типа задачи в базе.",
    "E4": "Записывать в историю каждую реальную попытку выполнения, синхронно с счётчиком попыток в задаче.",
    "E5": "Стабильные коды последней ошибки для мониторинга и инструкции по инцидентам.",
    "E6": "При повторе многошаговой задачи продолжать с последнего успешно выполненного шага.",
    "E7": "Учитывать лимиты запросов отдельно по видам операций (добавить канал, скачать сообщения и т.д.).",
    "E8": "Тест: после сбоя повтор не дублирует уже выполненные шаги пайплайна.",
    "F1": "Общая логика продюсеров: не создавать дубликаты задач и не переполнять очередь сверх целевого размера.",
    "F2": "Периодически создавать задачи переноса каналов, если нагрузка между аккаунтами разошлась больше чем на 5%.",
    "F3": "Отключить старый механизм переноса «в простое», когда работает балансировщик через очередь.",
    "F4": "Продюсер: ставить в очередь задачи сбора дополнительных данных по каналам, где они ещё не собраны.",
    "F5": "Продюсер: ставить в очередь обновление метаданных у каналов, которые давно не обновлялись.",
    "F6": "Выполнять задачу сбора дополнительных данных по каналу (скоринг, сообщения).",
    "F7": "Выполнять задачу обновления метаданных канала.",
    "F8": "Настроить расписание запуска продюсеров (cron или docker schedule).",
    "F9": "Включить в справочнике типы задач «обновление» и «сбор данных» после готовности адаптера.",
    "G1": "SQL-представления: сколько задач в каждом статусе и как давно ждёт самая старая.",
    "G2": "SQL-представления: процент свободного ресурса аккаунтов и список аккаунтов на паузе.",
    "G3": "HTTP API метрик очереди для админки и n8n.",
    "G4": "Алерты при росте очереди, большом числе отложенных задач или отсутствии свободных аккаунтов.",
    "G5": "Опционально автоматически повторять задачи, помеченные как застрявшие.",
    "G6": "Удалить старую SQLite-очередь массовых действий и воркер внутри main.",
    "G7": "Переписать тесты массовых async-операций на новую очередь в PostgreSQL.",
    "G8": "Синхронное массовое добавление и удаление каналов — только через PostgreSQL-очередь.",
    "G9": "Убрать дублирующий учёт лимитов запросов в памяти; оставить только учёт в базе.",
    "G10": "Документация runbook: как запускать воркер, продюсеры и переключатели режима.",
    "G11": "Пройти приёмочный чеклист MVP и полного соответствия ТЗ.",
}

# Оценка в идеальных часах (чистое время разработки, без встреч)
TASK_HOURS: dict[str, float] = {
    "Z1": 4, "Z2": 3, "Z3": 2, "Z4": 4,
    "A1": 2, "A2": 1, "A3": 4, "A4": 4, "A5": 6, "A6": 3, "A7": 3,
    "A8": 5, "A9": 2, "A10": 6, "A11": 4,
    "B1": 4, "B2": 3, "B3": 5, "B4": 6, "B5": 6, "B6": 8, "B7": 4,
    "B8": 5, "B9": 3, "B10": 12,
    "C1": 4, "C2": 8, "C3": 4, "C4": 6, "C5": 3, "C6": 4, "C7": 2,
    "C8": 8, "C9": 6,
    "D1": 6, "D2": 6, "D3": 8, "D4": 4, "D5": 3, "D6": 5, "D7": 6,
    "D8": 8, "D9": 6, "D10": 4, "D11": 1, "D12": 12,
    "E1": 4, "E2": 6, "E3": 4, "E4": 3, "E5": 3, "E6": 8, "E7": 6, "E8": 4,
    "F1": 4, "F2": 8, "F3": 1, "F4": 6, "F5": 5, "F6": 8, "F7": 5, "F8": 3, "F9": 2,
    "G1": 3, "G2": 3, "G3": 4, "G4": 6, "G5": 4, "G6": 6, "G7": 8, "G8": 6,
    "G9": 4, "G10": 4, "G11": 4,
}

block_why = {
    "Z": "Зафиксировать границы до кода: не смешать PG-очередь ТЗ с Huey-прототипом.",
    "A": "Diff §4: нет персистентной доменной очереди, типов задач, учёта ресурса в discovery.",
    "B": "Схема без репозиториев не работает; инкапсуляция SQL и инвариантов ТЗ.",
    "C": "ТЗ требует отдельный воркер с postpone; action_queue — только FIFO bulk.",
    "D": "Мост PG ↔ SessionClump; выполнение на зарезервированном аккаунте.",
    "E": "Связать classify_telethon_error с retry/postpone/failed задачи.",
    "F": "Продюсеры channel_balancer, collect_extra_data, update_channel по ТЗ §21–24.",
    "G": "Метрики, алерты, удаление legacy action_queue и дублей RPH.",
}

# Спринты: длительность в неделях, накопление
SPRINTS = [
    {"id": "S1", "name": "PG схема + repos", "blocks": "Z, A, B", "duration_w": 1.0, "task_ids": None, "mvp": True},
    {"id": "S2", "name": "Worker + методы clump", "blocks": "Z4, C, D1–D2", "duration_w": 1.0, "task_ids": None, "mvp": True},
    {"id": "S3", "name": "Адаптер + E2E MVP", "blocks": "D3–D12", "duration_w": 1.5, "task_ids": None, "mvp": True},
    {"id": "S4", "name": "Ошибки и идемпотентность", "blocks": "E", "duration_w": 1.0, "task_ids": None, "mvp": False},
    {"id": "S5", "name": "Продюсеры + мониторинг", "blocks": "F, G1–G5", "duration_w": 1.5, "task_ids": None, "mvp": False},
    {"id": "S6", "name": "Cleanup + приёмка", "blocks": "G6–G11", "duration_w": 0.5, "task_ids": None, "mvp": False},
]

MILESTONES = [
    ("M0", "Старт проекта", 0, PROJECT_START),
    ("M1", "PG + repos готовы (конец S1)", 1, None),
    ("M2", "Worker + dispatch (конец S2)", 2, None),
    ("M3", "MVP: add-channels через PG (конец S3)", 4, None),
    ("M4", "Retry + attempts (конец S4)", 5, None),
    ("M5", "Продюсеры + метрики (конец S5)", 7, None),
    ("M6", "Полное ТЗ принято (конец S6)", 8, None),
]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
MVP_FILL = PatternFill("solid", fgColor="C6EFCE")
NON_MVP_FILL = PatternFill("solid", fgColor="FFC7CE")
MILESTONE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GANTT_FILL = PatternFill("solid", fgColor="5B9BD5")


def _task_primary_sprint(tid: str) -> str:
    letter, num_s = tid[0], tid[1:]
    num = int(num_s)
    if letter == "Z":
        return "S1" if num <= 3 else "S2"
    if letter in ("A", "B"):
        return "S1"
    if letter == "C":
        return "S2"
    if letter == "D":
        return "S2" if num <= 2 else "S3"
    if letter == "E":
        return "S4"
    if letter == "F":
        return "S5"
    if letter == "G":
        return "S5" if num <= 5 else "S6"
    return "?"


def _week_to_date(week_offset: float) -> date:
    """week_offset=1 → конец 1-й недели (пятница)."""
    days = int(round(week_offset * 5))
    d = PROJECT_START
    added = 0
    while added < days:
        d += timedelta(days=1)
        if d.weekday() < 5:
            added += 1
    return d


def _sprint_windows() -> list[dict]:
    """Вычисляет start_w, end_w, даты для каждого спринта."""
    cur = 0.0
    out = []
    for sp in SPRINTS:
        start = cur
        end = cur + sp["duration_w"]
        tids = [t[1] for t in tasks if _task_primary_sprint(t[1]) == sp["id"]]
        out.append({
            **sp,
            "start_w": start,
            "end_w": end,
            "start_date": _week_to_date(start) if start > 0 else PROJECT_START,
            "end_date": _week_to_date(end),
            "task_count": len(tids),
            "task_ids": tids,
        })
        cur = end
    return out


def _style_header(ws, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _style_table(ws, min_row: int = 2) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER


def _add_waterfall_sheet(wb: Workbook, sprint_windows: list[dict]) -> None:
    ws = wb.create_sheet("Waterfall")
    ws["A1"] = "Waterfall: накопление срока и задач по этапам"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")

    headers = [
        "Этап", "Спринт", "Длит. (нед)", "База (скрыт.)", "Прирост (нед)",
        "Накоплено (нед)", "Задач за этап", "Накоплено задач", "MVP", "Веха",
    ]
    hr = 3
    for c, h in enumerate(headers, 1):
        ws.cell(hr, c, h)
    _style_header(ws, len(headers))

    data_start = hr + 1
    cum_tasks = 0
    row = data_start
    ws.cell(row, 1, "Старт")
    ws.cell(row, 4, 0)
    ws.cell(row, 5, 0)
    ws.cell(row, 6, 0)
    ws.cell(row, 7, 0)
    ws.cell(row, 8, 0)
    row += 1

    base_w = 0.0
    for sp in sprint_windows:
        cum_tasks += sp["task_count"]
        ws.cell(row, 1, sp["name"])
        ws.cell(row, 2, sp["id"])
        ws.cell(row, 3, sp["duration_w"])
        ws.cell(row, 4, base_w)
        ws.cell(row, 5, sp["duration_w"])
        ws.cell(row, 6, sp["end_w"])
        ws.cell(row, 7, sp["task_count"])
        ws.cell(row, 8, cum_tasks)
        ws.cell(row, 9, "Да" if sp["mvp"] else "Нет")
        if sp["id"] == "S3":
            ws.cell(row, 10, "★ MVP")
            for c in range(1, 11):
                ws.cell(row, c).fill = MILESTONE_FILL
        base_w = sp["end_w"]
        row += 1

    # Итого
    ws.cell(row, 1, "Полное ТЗ")
    ws.cell(row, 3, sum(s["duration_w"] for s in sprint_windows))
    ws.cell(row, 6, sprint_windows[-1]["end_w"])
    ws.cell(row, 7, len(tasks))
    ws.cell(row, 8, len(tasks))
    ws.cell(row, 10, "★ Приёмка")
    for c in range(1, 11):
        ws.cell(row, c).font = Font(bold=True)

    data_end = row
    _style_table(ws, data_start)

    # --- Waterfall chart (stacked bar: base + increment) ---
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "Waterfall: прирост длительности по спринтам (недели)"
    chart.y_axis.title = "Недели (накопительно)"
    chart.x_axis.title = "Этап"
    chart.height = 12
    chart.width = 22

    cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_end - 1)
    base_series = Reference(ws, min_col=4, min_row=hr, max_row=data_end - 1)
    inc_series = Reference(ws, min_col=5, min_row=hr, max_row=data_end - 1)
    chart.add_data(base_series, titles_from_data=True)
    chart.add_data(inc_series, titles_from_data=True)
    chart.set_categories(cats)
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = "FFFFFF"
        chart.series[0].graphicalProperties.line.solidFill = "FFFFFF"
        if len(chart.series) > 1:
            chart.series[1].graphicalProperties.solidFill = "4472C4"
    ws.add_chart(chart, f"A{data_end + 3}")

    # --- Cumulative tasks line chart ---
    line = LineChart()
    line.title = "Накопление закрытых задач"
    line.y_axis.title = "Задач"
    line.x_axis.title = "Этап"
    line.height = 10
    line.width = 22
    tasks_ref = Reference(ws, min_col=8, min_row=hr, max_row=data_end)
    line.add_data(tasks_ref, titles_from_data=True)
    line.set_categories(cats)
    if line.series:
        line.series[0].graphicalProperties.line.solidFill = "ED7D31"
        line.series[0].marker.symbol = "circle"
    ws.add_chart(line, f"A{data_end + 20}")

    widths = [28, 8, 10, 12, 12, 14, 12, 14, 6, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.cell(data_end + 2, 1, "Подсказка: синие столбцы — прирост недель; белая часть — уже пройденный срок.")
    ws.cell(data_end + 2, 1).font = Font(italic=True, color="666666")


def _task_hours(tid: str) -> float:
    if tid not in TASK_HOURS:
        raise KeyError(f"Нет оценки в часах для задачи {tid}")
    return TASK_HOURS[tid]


def _task_days_estimate(tid: str) -> float:
    return max(0.5, round(_task_hours(tid) / 6, 1))


def _days_to_story_points(days: float) -> float:
    if days <= 0.5:
        return 1.0
    if days <= 1.0:
        return 2.0
    if days <= 1.5:
        return 3.0
    if days <= 2.5:
        return 5.0
    if days <= 4.0:
        return 8.0
    return 13.0


def _task_vklad(mvp: str, blk: str) -> str:
    if mvp == "Да":
        return "must - база "
    if blk == "F":
        return "unic - фишка "
    if blk == "G":
        return "eco - экосистема"
    return "features - дополнение"


def _task_funkciya(tid: str, name: str) -> str:
    lower = name.lower()
    if lower.startswith(("система ", "должна ", "должен ")):
        title = name
    elif name.startswith("ADR"):
        title = f"Команда должна зафиксировать в ADR: {name[4:].strip()}"
    else:
        title = f"Система должна: {name}"
    plain = TASK_PLAIN_DESC.get(tid, "")
    if plain:
        return f"{title}\n\nПростыми словами: {plain}"
    return title


def _add_lidochka_dev_sheet(wb: Workbook, sprint_windows: list[dict]) -> None:
    """Лист в формате «Рассчётры по разработке» из Лидочка 2.0 (Разработка).xlsx."""
    ws = wb.create_sheet("Лидочка 2.0 (Разработка)")
    sp_by_id = {s["id"]: s for s in sprint_windows}
    sprint_results = {
        "S1": "PG схема + repos, тесты SQL",
        "S2": "Worker с mock adapter",
        "S3": "MVP: add-channels через PG",
        "S4": "Retry, attempts, идемпотентность",
        "S5": "Продюсеры + мониторинг",
        "S6": "Cleanup legacy, приёмка",
    }

    glossary = [
        "PG Queue Balancer — задачи по ТЗ Load Balancer for Telegram Channels",
        "Участник бизнес-процесса",
        "Функция / техническая задача для standalone_discovery",
        "1-Базовые функции\n 2-Функции, дифференцирующие продукт\n 3-Функции, вызывающие восторг\n 4-Экосистемные функции",
        "Эпик = блок Z–G плана внедрения",
        "Номер спринта S1–S6",
        "Календарь спринта (5-дневные недели)",
        "Ключевое событие / результат спринта",
        "Идеальные часы (без встреч)",
        'Оценка объёма (сила, сложность, риск) по шкале 1/2/3/5/8/13',
        "План реализации: зачем + критерий готовности",
        "Пояснения: зависимости, diff, MVP",
    ]
    ws.append(glossary)
    headers = [
        "Backlog",
        "Агент",
        "Функция",
        "Вклад (Must/Unic/Excite/Features/Eco)",
        "Метка /Эпик (по возможности)",
        "№ Спринта",
        "Спринт (14 дней)",
        "События",
        "Идеальные часы",
        "Объем\n (силия, сложность, риск)",
        "План реализации",
        "Пояснения",
    ]
    ws.append(headers)
    header_row = 2
    _style_header(ws, len(headers))

    data_start = 3
    for blk, tid, name, why, deps, mvp, done in tasks:
        bname, _phase, _sprint_meta = blocks_meta[blk]
        sid = _task_primary_sprint(tid)
        sp = sp_by_id[sid]
        hours = _task_hours(tid)
        days = _task_days_estimate(tid)
        volume = _days_to_story_points(days)
        sprint_label = f"{sp['start_date'].isoformat()} — {sp['end_date'].isoformat()}"
        plan = f"Зачем: {why}\n\nГотово когда: {done}\n\nID: {tid}"
        notes_parts = [f"MVP: {mvp}", f"Блок: {blk}"]
        if deps and deps != "—":
            notes_parts.append(f"Зависит от: {deps}")
        notes = "\n".join(notes_parts)

        ws.append([
            bname,
            "Система",
            _task_funkciya(tid, name),
            _task_vklad(mvp, blk),
            f"{blk} — {tid}",
            sid,
            sprint_label,
            sprint_results.get(sid, ""),
            hours,
            volume,
            plan,
            notes,
        ])

    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            if cell.value and str(cell.value).startswith("must"):
                cell.fill = MVP_FILL

    _style_table(ws, data_start)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A{header_row}:L{ws.max_row}"
    widths = [24, 10, 62, 22, 14, 10, 24, 28, 12, 12, 50, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_timeline_sheet(wb: Workbook, sprint_windows: list[dict]) -> None:
    ws = wb.create_sheet("Сроки")
    ws["A1"] = f"Детальный календарь (старт: {PROJECT_START.isoformat()}, 5-дневные недели)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:L1")

    # --- Блок 1: вехи ---
    r = 3
    ws.cell(r, 1, "Вехи")
    ws.cell(r, 1).font = Font(bold=True, size=12)
    r += 1
    mh = ["ID", "Веха", "Неделя", "Дата (ориентир)", "Задач к моменту", "Комментарий"]
    for c, h in enumerate(mh, 1):
        ws.cell(r, c, h)
    _style_header(ws, len(mh))
    r += 1
    cum = 0
    for mid, title, week, _ in MILESTONES:
        if mid == "M0":
            dt = PROJECT_START
            cum = 0
        else:
            dt = _week_to_date(week)
            cum = sum(sp["task_count"] for sp in sprint_windows if sp["end_w"] <= week)
            if mid == "M3":
                cum = sum(sp["task_count"] for sp in sprint_windows if sp["mvp"])
        comment = ""
        if mid == "M3":
            comment = "MVP ~48 задач"
        if mid == "M6":
            comment = "Полное ТЗ ~74 задачи"
            cum = len(tasks)
        ws.cell(r, 1, mid)
        ws.cell(r, 2, title)
        ws.cell(r, 3, week)
        ws.cell(r, 4, dt.isoformat())
        ws.cell(r, 5, cum)
        ws.cell(r, 6, comment)
        if mid in ("M3", "M6"):
            for c in range(1, 7):
                ws.cell(r, c).fill = MILESTONE_FILL
        r += 1

    # --- Блок 2: спринты ---
    r += 1
    ws.cell(r, 1, "Спринты (детализация)")
    ws.cell(r, 1).font = Font(bold=True, size=12)
    r += 1
    sh = [
        "Спринт", "Название", "Блоки", "Неделя нач.", "Неделя кон.",
        "Дата нач.", "Дата кон.", "Длит. (нед)", "Раб. дней", "Задач",
        "MVP", "Результат",
    ]
    for c, h in enumerate(sh, 1):
        ws.cell(r, c, h)
    _style_header(ws, len(sh))
    r += 1
    sprint_start_row = r
    for sp in sprint_windows:
        work_days = int(round(sp["duration_w"] * 5))
        ws.cell(r, 1, sp["id"])
        ws.cell(r, 2, sp["name"])
        ws.cell(r, 3, sp["blocks"])
        ws.cell(r, 4, sp["start_w"])
        ws.cell(r, 5, sp["end_w"])
        ws.cell(r, 6, sp["start_date"].isoformat())
        ws.cell(r, 7, sp["end_date"].isoformat())
        ws.cell(r, 8, sp["duration_w"])
        ws.cell(r, 9, work_days)
        ws.cell(r, 10, sp["task_count"])
        ws.cell(r, 11, "Да" if sp["mvp"] else "Нет")
        results = {
            "S1": "PG схема + repos, тесты SQL",
            "S2": "Worker с mock adapter",
            "S3": "MVP сквозной: add-channels через PG",
            "S4": "Retry, attempts, идемпотентность",
            "S5": "Продюсеры + мониторинг",
            "S6": "Cleanup legacy, приёмка",
        }
        ws.cell(r, 12, results.get(sp["id"], ""))
        if sp["mvp"]:
            ws.cell(r, 11).fill = MVP_FILL
        r += 1

    # --- Блок 3: Gantt по спринтам (недели 1–8) ---
    r += 1
    gantt_header_row = r
    ws.cell(r, 1, "Gantt по спринтам")
    ws.cell(r, 1).font = Font(bold=True, size=12)
    r += 1
    max_weeks = 8
    gh = ["Спринт", "Название"] + [f"Н{i}" for i in range(1, max_weeks + 1)]
    for c, h in enumerate(gh, 1):
        ws.cell(r, c, h)
    _style_header(ws, len(gh))
    r += 1
    for sp in sprint_windows:
        ws.cell(r, 1, sp["id"])
        ws.cell(r, 2, sp["name"])
        for w in range(1, max_weeks + 1):
            w_start = w - 1
            w_end = w
            if sp["end_w"] > w_start and sp["start_w"] < w_end:
                cell = ws.cell(r, 2 + w, "████")
                cell.fill = GANTT_FILL
                cell.font = Font(color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
        r += 1

    # --- Блок 4: оценка по блокам ---
    r += 1
    ws.cell(r, 1, "Сроки по блокам")
    ws.cell(r, 1).font = Font(bold=True, size=12)
    r += 1
    bh = ["Блок", "Название", "Фаза", "Спринт(ы)", "Задач", "MVP", "Оценка (нед)", "Комментарий"]
    for c, h in enumerate(bh, 1):
        ws.cell(r, c, h)
    _style_header(ws, len(bh))
    r += 1
    block_weeks = {
        "Z": 0.25, "A": 0.5, "B": 0.5, "C": 1.0,
        "D": 2.0, "E": 1.0, "F": 1.5, "G": 2.0,
    }
    for blk, (bname, phase, sprint) in blocks_meta.items():
        cnt = sum(1 for t in tasks if t[0] == blk)
        ws.cell(r, 1, blk)
        ws.cell(r, 2, bname)
        ws.cell(r, 3, phase)
        ws.cell(r, 4, sprint)
        ws.cell(r, 5, cnt)
        mvp = "Да" if blk in "ZABCD" else "Нет"
        ws.cell(r, 6, mvp)
        ws.cell(r, 7, block_weeks.get(blk, ""))
        ws.cell(r, 8, block_why.get(blk, ""))
        if mvp == "Да":
            ws.cell(r, 6).fill = MVP_FILL
        r += 1

    # --- Блок 5: задачи с датами ---
    r += 1
    ws.cell(r, 1, "Задачи: спринт и оценка дней")
    ws.cell(r, 1).font = Font(bold=True, size=12)
    r += 1
    th = [
        "ID", "Блок", "Спринт", "Неделя нач.", "Неделя кон.",
        "Дата нач.", "Дата кон.", "Часы", "Оценка (дн)", "MVP", "Задача",
    ]
    for c, h in enumerate(th, 1):
        ws.cell(r, c, h)
    _style_header(ws, len(th))
    r += 1
    task_start_row = r
    sp_by_id = {s["id"]: s for s in sprint_windows}
    for blk, tid, name, *_rest, mvp, _done in tasks:
        sid = _task_primary_sprint(tid)
        sp = sp_by_id[sid]
        hours = _task_hours(tid)
        days = _task_days_estimate(tid)
        ws.cell(r, 1, tid)
        ws.cell(r, 2, blk)
        ws.cell(r, 3, sid)
        ws.cell(r, 4, sp["start_w"])
        ws.cell(r, 5, sp["end_w"])
        ws.cell(r, 6, sp["start_date"].isoformat())
        ws.cell(r, 7, sp["end_date"].isoformat())
        ws.cell(r, 8, hours)
        ws.cell(r, 9, days)
        ws.cell(r, 10, mvp)
        ws.cell(r, 11, name)
        if mvp == "Да":
            ws.cell(r, 10).fill = MVP_FILL
        else:
            ws.cell(r, 10).fill = NON_MVP_FILL
        r += 1

    _style_table(ws, sprint_start_row)
    ws.freeze_panes = "A2"
    for i, w in enumerate([8, 22, 14, 10, 10, 12, 12, 8, 10, 8, 6, 40], 1):
        if i <= 12:
            ws.column_dimensions[get_column_letter(i)].width = min(w, 50)


def build_workbook() -> Workbook:
    wb = Workbook()
    sprint_windows = _sprint_windows()

    # --- Задачи ---
    ws = wb.active
    ws.title = "Задачи"
    headers = [
        "Блок", "Название блока", "ID", "Задача", "Зачем это нужно", "Deps",
        "Часы", "MVP", "Готово когда", "Фаза", "Спринт",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    for blk, tid, name, why, deps, mvp, done in tasks:
        bname, phase, sprint = blocks_meta[blk]
        ws.append([
            blk, bname, tid, name, why, deps, _task_hours(tid), mvp, done,
            phase, _task_primary_sprint(tid),
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):
        for cell in row:
            if cell.value == "Да":
                cell.fill = MVP_FILL
            elif cell.value == "Нет":
                cell.fill = NON_MVP_FILL
    _style_table(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{ws.max_row}"
    for i, w in enumerate([6, 22, 6, 40, 45, 18, 8, 8, 32, 8, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Сводка блоков ---
    ws2 = wb.create_sheet("Сводка блоков")
    h2 = ["Блок", "Название", "Фаза", "MVP", "Задач", "Часы", "Закрывает diff §15", "Зачем блок"]
    ws2.append(h2)
    _style_header(ws2, len(h2))
    for blk, (bname, phase, _sprint) in blocks_meta.items():
        cnt = sum(1 for t in tasks if t[0] == blk)
        hours = sum(_task_hours(t[1]) for t in tasks if t[0] == blk)
        diff_map = {
            "Z": "— (фундамент)", "A": "#1–6, #24–25 (частично)", "B": "#1–6, #7, #13, #16",
            "C": "#7, #9–12, #16–17, #32–34", "D": "#8, #10–12, #19, #23, #36–39",
            "E": "#13–15, #19, #22, #37", "F": "#24–28",
            "G": "#29–31, #35, legacy removal",
        }
        mvp = "Да" if blk in "ZABCD" else "Нет"
        ws2.append([blk, bname, phase, mvp, cnt, hours, diff_map[blk], block_why[blk]])
    total_hours = sum(TASK_HOURS.values())
    ws2.append(["", "ИТОГО", "", "", len(tasks), total_hours, "", ""])
    for cell in ws2[ws2.max_row]:
        cell.font = Font(bold=True)
    _style_table(ws2)
    ws2.freeze_panes = "A2"

    # --- Матрица diff ---
    ws3 = wb.create_sheet("Матрица diff")
    ws3.append(["Diff #", "Требование", "Закрывающие задачи"])
    _style_header(ws3, 3)
    for row in [
        [1, "task_types", "A3, A9, B2"],
        [2, "task_queue PG", "A5, B3–B5"],
        [3, "accounts по ТЗ", "A4, A10, B6–B7"],
        [4, "task_attempts", "A7, B9, E4"],
        [5, "account_resource_usage", "A6, B8, D5, G9"],
        [6, "Таблица каналов БД", "A8, D7, Z2"],
        [7, "Атомарный claim", "B4, C9"],
        [8, "Least-loaded", "C2, C5"],
        ["9–12", "Account / move / two accounts", "C4, D2, D4"],
        [13, "dedup_key", "A5, B3, F1"],
        [14, "Retry", "E1–E3"],
        ["15–16", "run_after, postpone/attempt", "B5, C3"],
        [17, "Приоритеты", "A9, C7"],
        [18, "1 задача = 1 работа", "D8, D9, G8"],
        [19, "Все попытки в лимите", "B8, D5, E7, G9"],
        [20, "min_available_resource_percent", "A9, C5"],
        [23, "Один канал — один аккаунт", "D7, D2"],
        ["24–28", "Продюсеры", "F2–F9"],
        ["29–31", "Мониторинг", "G1–G5"],
        [32, "1 задача на аккаунт", "A4, B6, C2"],
        [33, "dedup активных", "A5, B3"],
        [34, "Не стопорить очередь", "C3, C7"],
        [35, "Лимиты из БД", "A9, B2"],
        ["36–39", "Health, migrate, admin", "D6, D1"],
    ]:
        ws3.append(row)
    _style_table(ws3)

    # --- Спринты (расширенный) ---
    ws4 = wb.create_sheet("Спринты")
    ws4.append([
        "Спринт", "Название", "Длит. (нед)", "Неделя нач.", "Неделя кон.",
        "Дата нач.", "Дата кон.", "Задач", "Часы", "MVP", "Результат",
    ])
    _style_header(ws4, 11)
    results = {
        "S1": "PG схема + repos, тесты SQL",
        "S2": "Worker с mock adapter",
        "S3": "MVP сквозной: add-channels через PG",
        "S4": "Retry, attempts, идемпотентность",
        "S5": "Продюсеры + мониторинг",
        "S6": "Cleanup legacy, приёмка",
    }
    for sp in sprint_windows:
        sp_hours = sum(_task_hours(t[1]) for t in tasks if _task_primary_sprint(t[1]) == sp["id"])
        ws4.append([
            sp["id"], sp["name"], sp["duration_w"], sp["start_w"], sp["end_w"],
            sp["start_date"].isoformat(), sp["end_date"].isoformat(),
            sp["task_count"], sp_hours, "Да" if sp["mvp"] else "Нет",
            results.get(sp["id"], ""),
        ])
    _style_table(ws4)

    _add_waterfall_sheet(wb, sprint_windows)
    _add_timeline_sheet(wb, sprint_windows)
    _add_lidochka_dev_sheet(wb, sprint_windows)

    return wb


def build_lidochka_only_workbook(sprint_windows: list[dict] | None = None) -> Workbook:
    sprint_windows = sprint_windows or _sprint_windows()
    wb = Workbook()
    wb.remove(wb.active)
    _add_lidochka_dev_sheet(wb, sprint_windows)
    return wb


if __name__ == "__main__":
    sprint_windows = _sprint_windows()
    wb = build_workbook()
    wb.save(OUT)
    wb_l = build_lidochka_only_workbook(sprint_windows)
    wb_l.save(LIDOCHKA_OUT)
    print(f"Saved: {OUT} ({len(tasks)} tasks, {len(wb.sheetnames)} sheets)")
    print(f"Saved: {LIDOCHKA_OUT} (формат Лидочка 2.0)")
    print("Sheets:", ", ".join(wb.sheetnames))
